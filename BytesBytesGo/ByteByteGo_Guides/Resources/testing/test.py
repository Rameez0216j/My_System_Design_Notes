from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import pandas as pd
import os
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# Constants
URL = "https://blog.bytebytego.com/archive?sort=new"
EXCEL_FILE = "locked_non_ep_blogs.xlsx"

def setup_driver():
    options = Options()
    options.add_argument("--headless")  # Run in headless mode
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    return webdriver.Chrome(options=options)

def scroll_to_bottom(driver, pause_time=1):
    last_height = driver.execute_script("return document.body.scrollHeight")
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause_time)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height

def fetch_locked_non_ep_blogs():
    driver = setup_driver()
    driver.get(URL)
    scroll_to_bottom(driver)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    blog_entries = []

    containers = soup.find_all("div", class_="container-Qnseki")

    for container in containers:
        # Check for lock icon (premium content)
        if not container.find("svg", class_="lucide-lock"):
            continue

        # Extract blog <a> tag
        link_tag = container.find("a", attrs={"data-testid": "post-preview-title"})
        if link_tag:
            title = link_tag.text.strip()
            link = link_tag["href"].strip()
            if "EP" not in title:
                blog_entries.append((title, link))

    return blog_entries

def create_excel(blog_entries):
    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Blogs"

    # Add header row
    headers = ["Blog Name", "Blog Link", "Completed", "Remarks"]
    ws.append(headers)

    # Add data rows
    for entry in blog_entries:
        ws.append([entry[0], entry[1], False, ""])  # False for un-checked "Completed" by default

    # Apply checkbox validation for "Completed" column (C)
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', showDropDown=True)
    ws.add_data_validation(dv)
    for row in range(2, len(blog_entries) + 2):
        dv.add(ws[f'C{row}'])

    # Add conditional formatting: When "Completed" is checked, turn row green
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in range(2, len(blog_entries) + 2):
        # If the "Completed" checkbox (C column) is checked (True)
        ws[f'C{row}'].add_comment("Tick to mark as completed")  # Add a comment to guide the user
        ws.row_dimensions[row].height = 20  # Optional: Set row height

        # If checkbox is checked (True), highlight the entire row green
        ws.conditional_formatting.add(f"A{row}:D{row}", {
            "type": "expression",
            "formula": f'=$C{row}=TRUE',
            "format": green_fill
        })

    # Save workbook to file
    wb.save(EXCEL_FILE)
    print(f"✅ Excel file saved as {EXCEL_FILE}")

def main():
    blog_entries = fetch_locked_non_ep_blogs()
    create_excel(blog_entries)

if __name__ == "__main__":
    main()
