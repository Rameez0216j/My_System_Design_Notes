import os
import time
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# Config
URL = "https://blog.dailydoseofds.com/archive"
EXCEL_FILE = "Dailydoseofds_blogs.xlsx"
BLOG_DIV_CLASS = "container-Qnseki"

def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    return webdriver.Chrome(options=chrome_options)

def scroll_to_bottom(driver, pause_time=2, max_wait=20):
    last_height = driver.execute_script("return document.body.scrollHeight")
    wait_cycles = 0
    while wait_cycles < max_wait:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(pause_time)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            wait_cycles += 1
        else:
            wait_cycles = 0
        last_height = new_height

def wait_for_elements(driver, timeout=10):
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CLASS_NAME, BLOG_DIV_CLASS))
    )

def scrape_blogs():
    driver = setup_driver()
    driver.get(URL)
    scroll_to_bottom(driver)
    wait_for_elements(driver)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    blog_data = []
    for container in soup.find_all("div", class_=BLOG_DIV_CLASS):
        link_tag = container.find("a", {"data-testid": "post-preview-title"})
        if link_tag:
            name = link_tag.text.strip()
            link = link_tag.get("href").strip()
            blog_data.append((name, link))
    return blog_data

def load_existing_entries():
    if not os.path.exists(EXCEL_FILE):
        return set()
    df = pd.read_excel(EXCEL_FILE)
    return set(zip(df["Blog Name"], df["Link"]))

def save_to_excel(blog_entries):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Blog Name", "Link", "Completed", "Comments"])

    existing_entries = load_existing_entries()
    start_row = ws.max_row + 1
    new_entries = 0

    for name, link in blog_entries:
        if (name, link) not in existing_entries:
            ws.append([name, link, "FALSE", ""])
            new_entries += 1

    if new_entries > 0:
        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', showDropDown=True)
        ws.add_data_validation(dv)
        for row in range(start_row, start_row + new_entries):
            dv.add(ws[f'C{row}'])

            # Conditional formatting rule
            rule = FormulaRule(
                formula=[f'$C{row}=TRUE'],
                fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            )
            ws.conditional_formatting.add(f"A{row}:D{row}", rule)

    wb.save(EXCEL_FILE)
    print(f"✅ {new_entries} new blog(s) saved to {EXCEL_FILE}")

def main():
    print("🔍 Scraping blog...")
    blog_entries = scrape_blogs()
    print(f"📄 Total blogs scraped: {len(blog_entries)}")
    save_to_excel(blog_entries)

if __name__ == "__main__":
    main()
